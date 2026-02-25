"""Drawing file loader: filename parsing, filesystem scanning, thumbnail generation."""

import re
from pathlib import Path
from PIL import Image, ExifTags


# ─── Date helpers ─────────────────────────────────────────────────────────────

def _parse_year_to_date(raw: str | None) -> str | None:
    """
    Normalize catalog date strings to a sortable YYYY-MM-DD or YYYY string.
    Examples: "1966" → "1966-01-01", "c. 1999" → "1999-01-01", None → None
    """
    if not raw:
        return None
    raw = str(raw).strip()
    # Extract 4-digit year (possibly prefixed with "c.", "ca.", etc.)
    m = re.search(r'\b(\d{4})\b', raw)
    if m:
        return f"{m.group(1)}-01-01"
    return None


# Regex: YYYY-MM-DD[-description].ext
FILENAME_PATTERN = re.compile(
    r'^(\d{4}-\d{2}-\d{2})(?:-(.+?))?\.(jpe?g|png|gif|webp)$',
    re.IGNORECASE
)

SUPPORTED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}


def parse_filename(filename: str) -> dict:
    """
    Parse drawing metadata from filename.

    Examples:
      "2024-09-13-hobart st pittsburgh.jpeg"
        -> {drawn_date: "2024-09-13", title: "hobart st pittsburgh", file_ext: "jpeg"}
      "2016-05-06-trees.jpeg"
        -> {drawn_date: "2016-05-06", title: "trees", file_ext: "jpeg"}
      "2021-07-16.JPG"
        -> {drawn_date: "2021-07-16", title: None, file_ext: "jpg"}
    """
    match = FILENAME_PATTERN.match(filename)
    if match:
        return {
            'drawn_date': match.group(1),
            'title': match.group(2) if match.group(2) else None,
            'file_ext': match.group(3).lower()
        }
    # Fallback for non-standard filenames
    stem = Path(filename).stem
    ext = Path(filename).suffix.lstrip('.').lower()
    return {'drawn_date': None, 'title': stem, 'file_ext': ext}


def get_image_dimensions(filepath: str) -> tuple[int, int]:
    """Return (width, height) using PIL, respecting EXIF orientation."""
    try:
        with Image.open(filepath) as img:
            try:
                exif = img._getexif()
                if exif:
                    for tag, val in exif.items():
                        if ExifTags.TAGS.get(tag) == 'Orientation':
                            if val in (6, 8):  # 90 or 270 degrees
                                return img.size[1], img.size[0]
            except (AttributeError, KeyError):
                pass
            return img.size  # (width, height)
    except Exception:
        return (0, 0)


def generate_thumbnail(
    source_path: str,
    dest_path: str,
    size: tuple[int, int] = (400, 400)
) -> str:
    """
    Generate a JPEG thumbnail from source image.
    Handles EXIF rotation, RGBA->RGB conversion, and aspect ratio preservation.
    Adapted from art-journal storage.py.

    Returns dest_path on success.
    """
    Path(dest_path).parent.mkdir(parents=True, exist_ok=True)

    with Image.open(source_path) as img:
        # Handle EXIF rotation
        try:
            from PIL import ExifTags
            orientation_tag = None
            for tag_id, tag_name in ExifTags.TAGS.items():
                if tag_name == 'Orientation':
                    orientation_tag = tag_id
                    break

            if orientation_tag:
                exif = img._getexif()
                if exif is not None:
                    orientation_value = exif.get(orientation_tag)
                    if orientation_value == 3:
                        img = img.rotate(180, expand=True)
                    elif orientation_value == 6:
                        img = img.rotate(270, expand=True)
                    elif orientation_value == 8:
                        img = img.rotate(90, expand=True)
        except (AttributeError, KeyError, IndexError):
            pass

        # Convert to RGB (handles RGBA, P mode images)
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')

        img.thumbnail(size, Image.Resampling.LANCZOS)
        img.save(dest_path, format="JPEG", quality=85)

    return dest_path


def scan_user_dataset(dataset_path: str) -> list[dict]:
    """
    Walk a user's dataset directory and return list of drawing metadata dicts.
    Returns sorted by filename (chronological, since filenames start with YYYY-MM-DD).
    Does NOT touch the database.
    """
    results = []
    p = Path(dataset_path)

    if not p.exists() or not p.is_dir():
        return results

    for file in sorted(p.iterdir()):
        if not file.is_file():
            continue
        if file.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        if file.name.startswith('.'):  # skip .DS_Store etc
            continue

        meta = parse_filename(file.name)
        meta['filename'] = file.name
        meta['filepath'] = str(file.absolute())
        results.append(meta)

    return results


def scan_catalog_dataset(catalog_path: str, images_dir: str) -> list[dict]:
    """
    Load drawing metadata from an Excel catalog file.

    Expected columns (order doesn't matter, matched by header name):
      Image File, Title, Date, Medium, Location, Description (per PDF context)

    Returns list of dicts sorted by date, with same keys as scan_user_dataset().
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for catalog datasets: pip3 install openpyxl")

    wb = openpyxl.load_workbook(catalog_path)
    ws = wb.active

    # Build column index from header row
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    images_p = Path(images_dir)
    results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        filename = row[col['Image File'] - 1] if 'Image File' in col else None
        if not filename:
            continue

        filepath = images_p / filename
        if not filepath.exists():
            continue  # skip catalog entries without a matching image file

        raw_date = row[col['Date'] - 1] if 'Date' in col else None
        title = row[col['Title'] - 1] if 'Title' in col else None
        medium = row[col['Medium'] - 1] if 'Medium' in col else None
        location = row[col['Location'] - 1] if 'Location' in col else None
        description = row[col.get('Description (per PDF context)', 'Description') - 1] \
            if col.get('Description (per PDF context)') or col.get('Description') else None

        ext = Path(filename).suffix.lstrip('.').lower()

        results.append({
            'filename': filename,
            'filepath': str(filepath.absolute()),
            'drawn_date': _parse_year_to_date(raw_date),
            'title': str(title) if title else Path(filename).stem,
            'file_ext': ext,
            # Extra catalog fields stored as JSON-serialisable dict in analysis_json later
            'catalog_medium': str(medium) if medium else None,
            'catalog_location': str(location) if location else None,
            'catalog_description': str(description) if description else None,
            'catalog_raw_date': str(raw_date) if raw_date else None,
        })

    # Sort by drawn_date (None dates go to the end)
    results.sort(key=lambda r: (r['drawn_date'] is None, r['drawn_date'] or '', r['filename']))
    return results
